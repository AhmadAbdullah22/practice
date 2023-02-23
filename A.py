import csv
import openpyxl

with open("B.txt", 'r') as file:
    data = file.readlines()
print(data)


with open('B.txt') as C:
    data = csv.reader(C)
    for line in data:
        print(line)


wb = openpyxl.load_workbook('A.xlsx')
ws = wb.active

cell_value = ws['A1'].value
print(cell_value)

ws['B1'] = 'Abdullah'
wb.save('A.xlsx')
cell_value_2 = ws['B1'].value
print(cell_value_2)